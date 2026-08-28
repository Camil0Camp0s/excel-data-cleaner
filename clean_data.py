"""
Data Cleaner & Report Generator
--------------------------------
Toma un CSV/Excel "sucio" (típico export de un sistema de ventas) y entrega:
  1. Un Excel limpio y formateado, listo para usar.
  2. Una hoja de "Resumen" con métricas clave (ventas por ciudad, producto top, etc).
  3. Un log en consola de todo lo que se corrigió, para que el cliente vea el trabajo hecho.

Uso:
    python clean_data.py sample_messy_data.csv reporte_limpio.xlsx
"""

import sys
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def load_data(path: str) -> pd.DataFrame:
    if path.lower().endswith(".csv"):
        return pd.read_csv(path)
    return pd.read_excel(path)


def clean_text_columns(df: pd.DataFrame) -> pd.DataFrame:
    df["Nombre"] = df["Nombre"].str.strip().str.title()
    df["Email"] = df["Email"].str.strip().str.lower()
    df["Ciudad"] = df["Ciudad"].fillna("Desconocido").str.strip().str.title()
    df["Ciudad"] = df["Ciudad"].replace("", "Desconocido")
    df["Producto"] = df["Producto"].str.strip().str.title()
    return df


def normalize_dates(df: pd.DataFrame) -> pd.DataFrame:
    # Los datos traen fechas en varios formatos (DD/MM/YYYY, MM/DD/YYYY, YYYY-MM-DD, YYYY.MM.DD).
    # pandas no puede adivinar el formato correcto para todas a la vez, así que se
    # intenta cada formato conocido en orden y se completa con lo que sí matchee.
    formatos = ["%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%Y.%m.%d", "%d-%m-%Y"]
    parsed = pd.Series(pd.NaT, index=df.index)
    for fmt in formatos:
        mask = parsed.isna()
        parsed[mask] = pd.to_datetime(df.loc[mask, "Fecha_Compra"], format=fmt, errors="coerce")
    df["Fecha_Compra"] = parsed
    return df


def handle_missing_values(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    stats = {
        "precios_faltantes": int(df["Precio"].isna().sum()),
        "ciudades_desconocidas": int((df["Ciudad"] == "Desconocido").sum()),
    }
    # Precio faltante se rellena con el precio promedio de ese mismo producto
    df["Precio"] = df.groupby("Producto")["Precio"].transform(lambda s: s.fillna(s.mean()))
    return df, stats


def remove_duplicates(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    antes = len(df)
    df = df.drop_duplicates(subset=["Email", "Fecha_Compra", "Producto"])
    despues = len(df)
    return df, antes - despues


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    df["Total"] = df["Cantidad"] * df["Precio"]
    resumen_ciudad = df.groupby("Ciudad")["Total"].sum().sort_values(ascending=False)
    resumen_producto = df.groupby("Producto")["Cantidad"].sum().sort_values(ascending=False)

    filas = [
        ("Ventas totales", f"${df['Total'].sum():,.0f}"),
        ("Número de transacciones", len(df)),
        ("Ciudad con más ventas", resumen_ciudad.index[0]),
        ("Producto más vendido (unidades)", resumen_producto.index[0]),
        ("", ""),
        ("Ventas por ciudad", ""),
    ]
    for ciudad, total in resumen_ciudad.items():
        filas.append((ciudad, f"${total:,.0f}"))
    filas.append(("", ""))
    filas.append(("Unidades por producto", ""))
    for producto, unidades in resumen_producto.items():
        filas.append((producto, int(unidades)))

    return pd.DataFrame(filas, columns=["Métrica", "Valor"])


def style_sheet(ws):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length + 4


def main():
    if len(sys.argv) != 3:
        print("Uso: python clean_data.py <archivo_entrada> <archivo_salida.xlsx>")
        sys.exit(1)

    entrada, salida = sys.argv[1], sys.argv[2]

    df = load_data(entrada)
    filas_originales = len(df)

    df = clean_text_columns(df)
    df = normalize_dates(df)
    df, dup_removidas = remove_duplicates(df)
    df, missing_stats = handle_missing_values(df)
    df = df.sort_values("Fecha_Compra").reset_index(drop=True)

    resumen = build_summary(df)

    with pd.ExcelWriter(salida, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Datos Limpios", index=False)
        resumen.to_excel(writer, sheet_name="Resumen", index=False)
        for sheet in writer.sheets.values():
            style_sheet(sheet)

    print("=== Reporte de limpieza ===")
    print(f"Filas originales:        {filas_originales}")
    print(f"Duplicados eliminados:   {dup_removidas}")
    print(f"Precios faltantes rellenados (con promedio del producto): {missing_stats['precios_faltantes']}")
    print(f"Ciudades desconocidas marcadas: {missing_stats['ciudades_desconocidas']}")
    print(f"Filas finales:           {len(df)}")
    print(f"\nArchivo generado: {salida}")


if __name__ == "__main__":
    main()
